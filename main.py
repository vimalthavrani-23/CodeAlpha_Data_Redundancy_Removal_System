from openpyxl import Workbook, load_workbook
import os

FILE_NAME = "database.xlsx"

def create_database():
    if not os.path.exists(FILE_NAME):
        wb = Workbook()
        ws = wb.active
        ws.title = "Data"
        ws.append(["ID", "Name", "Email"])
        wb.save(FILE_NAME)

def add_data():
    wb = load_workbook(FILE_NAME)
    ws = wb["Data"]

    data_id = input("Enter ID: ")
    name = input("Enter Name: ")
    email = input("Enter Email: ")

    # Duplicate check
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] == data_id or row[2] == email:
            print("Duplicate data found! Data not added.")
            return

    ws.append([data_id, name, email])
    wb.save(FILE_NAME)
    print("Data added successfully.")

def view_data():
    wb = load_workbook(FILE_NAME)
    ws = wb["Data"]

    print("\nStored Data:\n")
    for row in ws.iter_rows(values_only=True):
        print(row)

create_database()

while True:
    print("\n===== Data Redundancy Removal System =====")
    print("1. Add Data")
    print("2. View Data")
    print("3. Exit")

    choice = input("Enter Choice: ")

    if choice == "1":
        add_data()
    elif choice == "2":
        view_data()
    elif choice == "3":
        print("Thank You")
        break
    else:
        print("Invalid Choice")
